import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    print('Adding new column with 10% discount of the original price:')
    for row in range(2, sheet.max_row + 1):
        # print(cell.value)
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    print('Adding new barchar with values of column 4, raws: 2,3,4:')

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')



    wb.save(filename)

# //////////////////////////////////////
# cell = sheet['a1']
# cell = sheet.cell(1,1)

# print(cell.value)
# print(sheet.max_row)

# https://openpyxl.readthedocs.io/en/3.1/index.html
# https://openpyxl.readthedocs.io/en/3.1/py-modindex.html

# //////////////////////////////////////